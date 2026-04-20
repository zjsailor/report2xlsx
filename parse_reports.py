# -*- coding: utf-8 -*-
"""
MD转Excel解析器 - 用于PGT-M胚胎植入前遗传学检测报告
将MD格式的报告解析并写入Excel表格

功能说明:
    1. 跳过预实验报告（包含`_PGTMF_`的文件）
    2. 提取目标变异/SNP分型一致性到U列，值包括：一致、不一致、不一致（位点扩增ADO）

用法:
    python parse_reports.py [-i <md文件夹>] [-o <输出excel>]
    python parse_reports.py -i D:\md2excel\md_output -o D:\md2excel\Info.xlsx

依赖:
    pip install openpyxl
"""

import re
import sys
import os
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")


def clean_html(text):
    """清理HTML标签和多余空白"""
    if not text:
        return ""
    text = re.sub(r"<br\s*/?>", " ", text)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\s+", " ", text)
    text = text.strip("；;,。. ")
    return text.strip()


def extract_age_from_line(line):
    """从行中提取年龄 - 处理多种格式"""
    # 格式1: "年 龄 32" (with spaces)
    m = re.search(r"年\s*龄\s*(\d+)", line)
    if m:
        return m.group(1)
    # 格式2: "年龄:32" or "年龄 32"
    m = re.search(r"年龄[：:\s]+(\d+)", line)
    if m:
        return m.group(1)
    return ""


def extract_target_mutation_results(content):
    """提取目标变异检测结果 - 返回 (mutation_results, snp_results, target_gene)

    SNP分型一致性(snp_consistency)可能的值:
        - 一致
        - 不一致
        - 不一致（位点扩增ADO）
        - 不一致（位点扩增）
        - 不一致（检测异常）
        - 不一致（母源/父源污染等原因）
    """
    mutation_results = {}
    snp_results = {}
    target_gene = ""

    lines = content.split("\n")
    in_target_section = False
    sample_name_col = -1

    for i, line in enumerate(lines):
        stripped = line.strip()

        if "目标变异检测结果" in stripped or "目标变异/SNP分型" in stripped:
            in_target_section = True
            continue

        if not in_target_section:
            continue

        if not stripped or stripped.startswith("|---"):
            continue

        if any(
            x in stripped
            for x in [
                "SNP可用位点",
                "SNP单体型",
                "位点验证",
                "附件",
                "结果说明",
                "图谱",
            ]
        ):
            break

        if re.match(r"^[A-Z][A-Za-z0-9_]+,", stripped):
            target_gene = stripped.split(",")[0]
            continue

        if "样本名称" in stripped and "目标变异" in stripped:
            parts = [p.strip() for p in stripped.split("|") if p.strip()]
            for j, p in enumerate(parts):
                if "样本名称" in p:
                    sample_name_col = j
                    break
            continue

        if sample_name_col >= 0 and "|" in stripped:
            parts = [p.strip() for p in stripped.split("|") if p.strip()]
            if len(parts) > sample_name_col:
                sample_id = parts[sample_name_col]
                if (
                    sample_id
                    and sample_id not in ["样本名称", target_gene]
                    and not sample_id.startswith("基因")
                ):
                    if len(parts) > sample_name_col + 1:
                        mutation1 = (
                            parts[sample_name_col + 1]
                            if len(parts) > sample_name_col + 1
                            else ""
                        )
                    else:
                        mutation1 = ""
                    snp_consistency = ""
                    for p in reversed(parts):
                        if p in [
                            "一致",
                            "不一致",
                            "不一致（位点扩增ADO）",
                            "不一致（位点扩增）",
                            "不一致（检测异常）",
                        ]:
                            snp_consistency = p
                            break
                        # 处理空白或"-"的情况（染色体交换重组、母源污染等导致无法判断）
                        if p == "-" or p == "—" or p == "":
                            if snp_consistency == "":
                                snp_consistency = "不一致"
                    mutation_results[sample_id] = (mutation1, "")
                    snp_results[sample_id] = snp_consistency
                    continue

        parts = stripped.split()
        if len(parts) >= 2:
            first_part = parts[0]
            if (
                first_part
                and not first_part.startswith("基因")
                and not first_part.startswith("DMD")
                and not first_part.startswith("FBN")
            ):
                if first_part not in [
                    "SNP连锁分析判断结果",
                    "与突变位点检测结果",
                    "是否一致",
                    "可用位点数",
                ]:
                    mutation1 = ""
                    snp_consistency = ""
                    if len(parts) >= 2:
                        mutation1 = parts[1]
                    for p in parts[2:]:
                        if p in [
                            "一致",
                            "不一致",
                            "不一致（位点扩增ADO）",
                            "不一致（位点扩增）",
                        ]:
                            snp_consistency = p
                            break
                    mutation_results[first_part] = (mutation1, "")
                    snp_results[first_part] = snp_consistency

    return mutation_results, snp_results, target_gene


def find_mutation_and_snp_by_partial_id(embryo_id, mutation_results, snp_results):
    """根据胚胎ID在mutation_results和snp_results中查找匹配

    胚胎ID可能存在格式差异，如:
        - 胚胎表格: "241228CC Y_E2" (有空格)
        - SNP表格: "241228CCY_E2" (无空格)
    匹配时优先精确匹配，失败后使用前缀heuristic匹配。
    """
    if not embryo_id:
        return ("", ""), ""

    if embryo_id in mutation_results:
        return mutation_results[embryo_id], snp_results.get(embryo_id, "")

    clean_id = embryo_id.replace(" ", "").replace("_", "")

    # 优先：精确匹配（去除所有空格后匹配）
    for key in mutation_results.keys():
        if not key:
            continue
        key_clean = key.replace(" ", "").replace("_", "")
        if key_clean == clean_id:
            return mutation_results[key], snp_results.get(key, "")

    # 次优：基于数字+前缀的heuristic匹配
    import re

    embryo_nums = re.findall(r"\d+", embryo_id)
    if embryo_nums:
        embryo_first_num = embryo_nums[0]
        embryo_prefix = re.split(r"\d+", embryo_id)[0] if embryo_id else ""

        for key in mutation_results.keys():
            if not key:
                continue
            key_clean = key.replace(" ", "").replace("_", "")
            if key_clean == clean_id:
                continue  # already handled above

            key_nums = re.findall(r"\d+", key)
            if key_nums and key_nums[0] == embryo_first_num:
                key_prefix = re.split(r"\d+", key)[0] if key else ""
                if embryo_prefix and key_prefix:
                    if (
                        embryo_prefix in key
                        or key in embryo_prefix
                        or (
                            len(embryo_prefix) >= 2
                            and len(key_prefix) >= 2
                            and embryo_prefix[:2] == key_prefix[:2]
                        )
                    ):
                        return mutation_results[key], snp_results.get(key, "")

    # 最后：基于前缀的匹配（8字符起始）
    if clean_id and len(clean_id) >= 8:
        for key in mutation_results.keys():
            if not key:
                continue
            key_clean = key.replace(" ", "").replace("_", "")
            if key_clean == clean_id:
                continue
            if key.startswith(embryo_id[:8]) or embryo_id.startswith(key[:8]):
                return mutation_results[key], snp_results.get(key, "")

    return ("", ""), ""

    # 尝试精确匹配
    if embryo_id in mutation_results:
        return mutation_results[embryo_id], snp_results.get(embryo_id, "")

    # 规范化：去除所有空格
    clean_id = embryo_id.replace(" ", "").replace("_", "")

    # 遍历查找最佳匹配
    for key in mutation_results.keys():
        if not key:
            continue
        key_clean = key.replace(" ", "").replace("_", "")

        # 完全去除空格后匹配
        if key_clean == clean_id:
            return mutation_results[key], snp_results.get(key, "")

        # 提取数字部分进行匹配（处理 "241228CC Y_E2" 匹配 "241228CCY_E2" 的情况）
        import re

        embryo_nums = re.findall(r"\d+", embryo_id)
        key_nums = re.findall(r"\d+", key)
        if embryo_nums and key_nums:
            # 如果数字部分匹配且前缀/后缀有重叠
            if embryo_nums[0] == key_nums[0]:  # 第一个数字串相同（如241228CC）
                embryo_prefix = (
                    re.split(r"\d+", embryo_id)[0]
                    if re.split(r"\d+", embryo_id)
                    else ""
                )
                key_prefix = re.split(r"\d+", key)[0] if re.split(r"\d+", key) else ""
                # 前缀相似（如CC, CCY）
                if embryo_prefix and key_prefix:
                    # 检查是否一个包含另一个或相似
                    if (
                        embryo_prefix in key
                        or key in embryo_prefix
                        or embryo_prefix[:2] == key_prefix[:2]
                        if len(embryo_prefix) > 1 and len(key_prefix) > 1
                        else False
                    ):
                        return mutation_results[key], snp_results.get(key, "")

        # 基于前缀的匹配（8字符起始）
        if clean_id and len(clean_id) >= 8:
            if key.startswith(embryo_id[:8]) or embryo_id.startswith(key[:8]):
                return mutation_results[key], snp_results.get(key, "")

    return ("", ""), ""


def parse_md_file(filepath, patient_name):
    """解析单个MD文件"""
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    result = {
        "patient": patient_name,
        "female_name": "",
        "male_name": "",
        "female_age": "",
        "male_age": "",
        "sample_date": "",
        "report_no": "",
        "sample_barcode": "",
        "disease": "",
        "gene": "",
        "mutation1": "",
        "mutation2": "",
        "embryos": [],
    }

    # 提取目标变异检测结果
    target_mutations, target_snp, target_gene = extract_target_mutation_results(content)

    lines = content.split("\n")

    if "南宁市第二人民医院" in content:
        # 苏肖榕 format from 南宁市第二人民医院
        for line in lines:
            # Extract female name
            if "女方姓名" in line:
                parts = [p.strip() for p in line.split("|")]
                parts = [p for p in parts if p]
                for i, p in enumerate(parts):
                    if "女方姓名" in p and i + 1 < len(parts):
                        val = parts[i + 1]
                        if "苏肖榕" in val:
                            result["female_name"] = "苏肖榕"
                        elif "年龄" not in val:
                            result["female_name"] = val
                        break
            # Extract male name
            if "男方姓名" in line:
                parts = [p.strip() for p in line.split("|")]
                parts = [p for p in parts if p]
                for i, p in enumerate(parts):
                    if "男方姓名" in p and i + 1 < len(parts):
                        val = parts[i + 1]
                        if "黄强" in val:
                            result["male_name"] = "黄强"
                        elif "年龄" not in val:
                            result["male_name"] = val
                        break
            # Extract ages
            if "年龄" in line:
                ages = re.findall(r"(\d+)", line)
                if len(ages) >= 2:
                    result["female_age"] = ages[0]
                    result["male_age"] = ages[1]
            # Extract sample date
            if "收样日期" in line:
                m = re.search(r"收样日期[：:\s|]+(\d{4}-\d{2}-\d{2})", line)
                if m:
                    result["sample_date"] = m.group(1)

        # Extract gene and disease
        gene_m = re.search(r"检测基因名称[：:\s]+(\S+)[；|\s]", content)
        disease_m = re.search(r"检测疾病名称[：:\s]+([^(（]+)", content)
        mutation_m = re.search(r"位置[：:\s]+([^（]+)[（](母源|父源)[）]", content)

        if gene_m:
            result["gene"] = gene_m.group(1)
        if disease_m:
            result["disease"] = clean_html(disease_m.group(1))
        if mutation_m:
            result["mutation1"] = f"{mutation_m.group(1)}({mutation_m.group(2)})"

        # Parse embryo data
        start_parsing = False
        for line in lines:
            stripped = line.strip()
            if "胚胎编号" in stripped and stripped.startswith("|"):
                start_parsing = True
                continue
            if start_parsing and stripped.startswith("|"):
                if "---" in stripped or not stripped:
                    continue
                if any(
                    x in stripped
                    for x in [
                        "注释",
                        "结果分析",
                        "附件",
                        "申明",
                        "局限性",
                        "SNP可用",
                        "SNP单体型",
                    ]
                ):
                    break
                parts = [p.strip() for p in stripped.split("|")]
                parts = [p for p in parts if p]
                if len(parts) >= 3 and parts[0].startswith("T_"):
                    embryo_id = parts[0]
                    cnv = parts[1].strip() if len(parts) > 1 else ""
                    carrier = parts[-1].strip() if parts else ""
                    if embryo_id not in [e["id"] for e in result["embryos"]]:
                        embryo = {
                            "id": embryo_id,
                            "morphology": "",
                            "cnv": cnv,
                            "cnv_explain": "",
                            "aneuploidy": "",
                            "carrier_status": carrier,
                            "mutation_detection1": "",
                            "mutation_detection2": "",
                            "snp_consistency": "",
                        }
                        mutations, snp = find_mutation_and_snp_by_partial_id(
                            embryo_id, target_mutations, target_snp
                        )
                        if mutations:
                            embryo["mutation_detection1"] = mutations[0]
                            embryo["mutation_detection2"] = mutations[1]
                        embryo["snp_consistency"] = snp
                        result["embryos"].append(embryo)
    else:
        # Standard format
        for line in lines:
            # Extract female name
            if "女方姓名" in line:
                parts = [p.strip() for p in line.split("|")]
                parts = [p for p in parts if p]
                for i, p in enumerate(parts):
                    if "女方姓名" in p and i + 1 < len(parts):
                        val = parts[i + 1]
                        if (
                            val
                            and "女方" not in val
                            and "年龄" not in val
                            and "男方" not in val
                        ):
                            result["female_name"] = val
                            break
            # Extract male name
            if "男方姓名" in line:
                parts = [p.strip() for p in line.split("|")]
                parts = [p for p in parts if p]
                for i, p in enumerate(parts):
                    if "男方姓名" in p and i + 1 < len(parts):
                        val = parts[i + 1]
                        if (
                            val
                            and "男方" not in val
                            and "年龄" not in val
                            and "女方" not in val
                        ):
                            result["male_name"] = val
                            break
            # Extract ages - improved pattern to handle "年 龄 32" format
            if "年" in line and "龄" in line:
                age = extract_age_from_line(line)
                if age:
                    # 判断是女方还是男方
                    if "女方" in line:
                        result["female_age"] = age
                    if "男方" in line:
                        result["male_age"] = age
            # Extract sample date
            if "收样日期" in line:
                parts = [p.strip() for p in line.split("|")]
                parts = [p for p in parts if p]
                for i, p in enumerate(parts):
                    if "收样日期" in p and i + 1 < len(parts):
                        val = parts[i + 1]
                        if val and "收样日期" not in val:
                            result["sample_date"] = val
                            break
            # Extract report number
            if "送检编号" in line:
                parts = [p.strip() for p in line.split("|")]
                parts = [p for p in parts if p]
                for i, p in enumerate(parts):
                    if "送检编号" in p and i + 1 < len(parts):
                        val = parts[i + 1]
                        if val and "送检编号" not in val and "送检条码" not in val:
                            result["report_no"] = val
                            break
            # Extract sample barcode
            if "送检条码" in line:
                parts = [p.strip() for p in line.split("|")]
                parts = [p for p in parts if p]
                for i, p in enumerate(parts):
                    if "送检条码" in p and i + 1 < len(parts):
                        val = parts[i + 1]
                        if val and "送检条码" not in val and "送检编号" not in val:
                            result["sample_barcode"] = val
                            break

        # Extract genes
        gene_matches = re.findall(r"基因名称[：:\s]+([A-Za-z0-9_]+)", content)
        unique_genes = list(dict.fromkeys(gene_matches))
        if unique_genes:
            result["gene"] = ",".join(unique_genes[:2])

        # Extract diseases
        disease_matches = re.findall(
            r"疾病名称\([^)]+\)\d*[：:]\s*([^{\n|<]+)", content
        )
        unique_diseases = list(dict.fromkeys(disease_matches))
        if unique_diseases:
            result["disease"] = "；".join([clean_html(d) for d in unique_diseases[:2]])
        else:
            disease_match2 = re.search(r"疾病名称[：:]\s*([^{\n|<]+)", content)
            if disease_match2:
                result["disease"] = clean_html(disease_match2.group(1))

        # Extract mutations - look for patterns like "c.IVS1-14(G>A)(-28)（父源）" or "Exon2dup（母源）"
        mutation_regexes = [
            r"(c\.[A-Za-z0-9><_-]+(?:\([^)]*\))*)[（(](?:母源|父源)[)）]",
            r"(Exon\d+[-\d]*[a-zA-Z]*)[（(](?:母源|父源)[)）]",
        ]
        mutations_set = set()
        for regex in mutation_regexes:
            for m in re.finditer(regex, content):
                full_match = m.group(1)
                if "c." in full_match or "Exon" in full_match:
                    source_match = re.search(r"[（(](母源|父源)[)）]", m.group())
                    if source_match:
                        mutations_set.add(f"{full_match}({source_match.group(1)})")
        mutations = list(mutations_set)
        if mutations:
            result["mutation1"] = mutations[0] if len(mutations) >= 1 else ""
            result["mutation2"] = mutations[1] if len(mutations) >= 2 else ""

        embryos = []
        in_embryo_section = False
        header_col_map = {}
        current_line_idx = 0
        seen_embryo_ids = set()

        for i, line in enumerate(lines):
            stripped = line.strip()

            if (
                stripped.startswith("|检测结果|")
                or stripped.startswith("|结果信息|")
                or stripped.startswith("|胚胎编号|")
            ):
                in_embryo_section = True
                current_line_idx = 0
                continue

            if not in_embryo_section:
                continue

            if current_line_idx == 0:
                if (
                    "样本名称" in stripped
                    or "胚胎编号" in stripped
                    or "胚胎评级" in stripped
                    or "形态学" in stripped
                    or "检测结果" in stripped
                    or "评级" in stripped
                ):
                    parts = [p.strip() for p in stripped.split("|")]
                    header_col_map.clear()
                    for j, p in enumerate(parts):
                        if not p:
                            continue
                        p_lower = p.lower()
                        if "样本名称" in p or "胚胎编号" in p:
                            header_col_map["sample_name"] = j
                        if "形态学" in p or "评级" in p:
                            header_col_map["morphology"] = j
                        if (
                            "cnv" in p_lower
                            and "检测结果" not in p
                            and "异倍体" not in p
                        ):
                            header_col_map["cnv"] = j
                        if "解释" in p or "结果解释" in p:
                            header_col_map["cnv_explain"] = j
                        if "异倍体" in p:
                            header_col_map["aneuploidy"] = j
                        if "携带状态" in p or ("携带" in p and "基因突变" in p):
                            header_col_map["carrier"] = j
                    current_line_idx = 1
                    continue
                if (
                    stripped.startswith("|")
                    and "|" in stripped
                    and len(stripped.split("|")) > 3
                ):
                    parts = [p.strip() for p in stripped.split("|")]
                    potential_id = parts[1] if len(parts) > 1 else ""
                    if (
                        potential_id
                        and not any(c in potential_id for c in "（）：")
                        and len(potential_id) < 20
                    ):
                        current_line_idx = 1
                continue

            if any(
                x in stripped
                for x in [
                    "结果说明",
                    "附件",
                    "检测局限性",
                    "SNP可用位点统计",
                    "SNP单体型分型图谱",
                    "位点验证",
                    "目标变异检测结果",
                    "图谱",
                    "注释",
                    "SNP分型",
                    "家系基因",
                    "突变位点信息",
                    "二、",
                    "三、",
                    "附件一",
                    "附件二",
                ]
            ):
                break

            if not stripped or stripped.startswith("|---"):
                continue

            if stripped.startswith("##"):
                break

            if not stripped.startswith("|"):
                current_line_idx = 0
                continue

            parts = [p.strip() for p in stripped.split("|")]
            if not any(parts):
                continue

            sample_col = header_col_map.get("sample_name", 0)
            embryo_id = ""
            if sample_col < len(parts) and parts[sample_col]:
                embryo_id = parts[sample_col]

            if not embryo_id and len(parts) > 1 and parts[1]:
                embryo_id = parts[1]

            if embryo_id and embryo_id not in [
                "样本名称",
                "胚胎编号",
                "检测结果",
                "基因突变位点及",
                "SNP检测结果",
            ]:
                if embryo_id in seen_embryo_ids:
                    continue
                if any(c in embryo_id for c in "（）：") or len(embryo_id) > 25:
                    continue
                if any(
                    x in embryo_id
                    for x in ["基因", "位置", "上游", "下游", "检测结果", "样本名称"]
                ):
                    continue
                if "玉伟 _" in embryo_id:
                    continue
                seen_embryo_ids.add(embryo_id)

                embryo = {
                    "id": embryo_id,
                    "morphology": "",
                    "cnv": "",
                    "cnv_explain": "",
                    "aneuploidy": "",
                    "carrier_status": "",
                    "mutation_detection1": "",
                    "mutation_detection2": "",
                    "snp_consistency": "",
                }

                if "morphology" in header_col_map and header_col_map[
                    "morphology"
                ] < len(parts):
                    embryo["morphology"] = parts[header_col_map["morphology"]]
                if "cnv" in header_col_map and header_col_map["cnv"] < len(parts):
                    embryo["cnv"] = parts[header_col_map["cnv"]]
                if "cnv_explain" in header_col_map and header_col_map[
                    "cnv_explain"
                ] < len(parts):
                    embryo["cnv_explain"] = parts[header_col_map["cnv_explain"]]
                if "aneuploidy" in header_col_map and header_col_map[
                    "aneuploidy"
                ] < len(parts):
                    embryo["aneuploidy"] = parts[header_col_map["aneuploidy"]]
                if "carrier" in header_col_map and header_col_map["carrier"] < len(
                    parts
                ):
                    embryo["carrier_status"] = parts[header_col_map["carrier"]]

                mutations, snp = find_mutation_and_snp_by_partial_id(
                    embryo_id, target_mutations, target_snp
                )
                if mutations:
                    embryo["mutation_detection1"] = mutations[0]
                    embryo["mutation_detection2"] = mutations[1]
                embryo["snp_consistency"] = snp

                embryos.append(embryo)

        result["embryos"] = embryos

    return result


def get_all_md_files(md_folder):
    """获取md_folder下所有md文件，自动匹配患者名

    跳过规则:
        - 预实验报告（文件名包含`_PGTMF_`或`PGTMF_`）
    """
    md_folder = Path(md_folder)
    md_files = list(md_folder.glob("*.md"))

    file_mapping = []
    for md_file in md_files:
        # 跳过预实验报告（包含_PGTMF_的文件）
        if "_PGTMF_" in md_file.name or "PGTMF_" in md_file.name:
            continue
        # 提取患者名（去掉后缀）
        patient_name = md_file.stem
        # 去掉常见的报告后缀
        for suffix in [
            "_YKSZ_PGTM_240515_17C_000120240531150943461",
            "_YKSZ_PGTM_250421_17C_0001",
            "_YKSZ_PGTM_250106_32A_0003_非地贫",
            "_YKSZ_PGTM_250104_16G_0002",
            "_YKSZ_PGTM_250102_21C_0001",
            "-YKSZ_PGTM_220704_17C_01",
            "_YKSZ_PGTM_250206_18G_0002",
            "_YKSZ_PGTM_250107_24C_0001",
            "_250102_21C_0001",
            "_250421_17C_0001",
            "_240515_17C_000120240531150943461",
        ]:
            if patient_name.endswith(suffix):
                patient_name = patient_name[: -len(suffix)]
                break
        # 清理末尾的下划线和空格
        patient_name = patient_name.strip().rstrip("_").rstrip("-")
        file_mapping.append((md_file.name, patient_name))

    return file_mapping


def get_existing_barcodes(ws):
    """获取Excel中已有的送检条码集合"""
    existing_barcodes = set()
    for row in range(2, ws.max_row + 1):
        barcode = ws.cell(row, 3).value  # 送检条码在第3列
        if barcode:
            existing_barcodes.add(str(barcode).strip())
    return existing_barcodes


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="解析PGT-M胚胎检测报告MD文件并写入Excel"
    )
    parser.add_argument(
        "-i",
        "--input",
        default=r"D:\md2excel\md_output",
        help="MD文件所在文件夹 (默认: D:\\md2excel\\md_output)",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=r"D:\md2excel\Info.xlsx",
        help="输出Excel文件路径 (默认: D:\\md2excel\\Info.xlsx)",
    )
    parser.add_argument(
        "-m",
        "--mode",
        choices=["overwrite", "append"],
        default="overwrite",
        help="写入模式: overwrite=覆盖(默认), append=追加",
    )
    args = parser.parse_args()

    md_folder = args.input
    excel_path = args.output
    mode = args.mode

    # 获取所有MD文件
    files_to_process = get_all_md_files(md_folder)
    print(f"找到 {len(files_to_process)} 个MD文件")
    print(f"写入模式: {'追加' if mode == 'append' else '覆盖'}")

    # 检查Excel文件是否存在
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        if mode == "overwrite":
            # 覆盖模式：删除除表头外的所有行
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            print(f"已重置Excel，保留表头")
            existing_barcodes = set()
        else:
            # 追加模式：获取已有的送检条码
            existing_barcodes = get_existing_barcodes(ws)
            print(f"已有 {len(existing_barcodes)} 条记录，将跳过重复项")
    else:
        # 创建新Excel并添加表头
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [
            "文件名",
            "送检编号",
            "送检条码",
            "收样日期",
            "女方姓名",
            "女方年龄",
            "男方姓名",
            "男方年龄",
            "疾病",
            "基因",
            "突变1",
            "突变2",
            "胚胎编号",
            "形态学评级",
            "CNV检测结果",
            "CNV结果解释",
            "异倍体检测结果",
            "携带状态",
            "目标变异检测结果",
            "变异2相关信息",
            "目标变异/SNP分型一致性",
        ]
        ws.append(headers)
        print(f"创建新Excel，已添加表头")
        existing_barcodes = set()

    new_rows_count = 0
    skipped_rows_count = 0

    for filename, patient_name in files_to_process:
        print(f"\n处理: {filename}...")
        try:
            data = parse_md_file(f"{md_folder}\\{filename}", patient_name)

            print(f"  女性: {data['female_name']}, 男性: {data['male_name']}")
            print(f"  女年龄: {data['female_age']}, 男年龄: {data['male_age']}")
            print(f"  收样日期: {data['sample_date']}")
            print(f"  送检条码: {data['sample_barcode']}")
            print(f"  疾病: {data['disease'][:50] if data['disease'] else 'N/A'}...")
            print(f"  基因: {data['gene']}, 突变: {data['mutation1']}")
            print(f"  发现 {len(data['embryos'])} 个胚胎")

            # 追加模式：用送检条码判断是否已处理
            if mode == "append" and data["sample_barcode"]:
                barcode = str(data["sample_barcode"]).strip()
                if barcode in existing_barcodes:
                    print(f"  跳过（送检条码已存在）: {barcode}")
                    skipped_rows_count += len(data["embryos"])
                    continue

            for embryo in data["embryos"]:
                row_data = [
                    filename.replace(".md", ".pdf"),  # Col 1
                    data["report_no"],  # Col 2
                    data["sample_barcode"],  # Col 3
                    data["sample_date"],  # Col 4
                    data["female_name"],  # Col 5
                    data["female_age"],  # Col 6
                    data["male_name"],  # Col 7
                    data["male_age"],  # Col 8
                    data["disease"],  # Col 9
                    data["gene"],  # Col 10
                    data["mutation1"],  # Col 11
                    data["mutation2"],  # Col 12
                    embryo["id"],  # Col 13
                    embryo["morphology"],  # Col 14
                    embryo["cnv"],  # Col 15
                    embryo["cnv_explain"],  # Col 16
                    embryo["aneuploidy"],  # Col 17
                    embryo["carrier_status"],  # Col 18
                    embryo.get("mutation_detection1", ""),  # Col 19
                    embryo.get("mutation_detection2", ""),  # Col 20
                    embryo.get("snp_consistency", ""),  # Col 21
                ]
                ws.append(row_data)
                new_rows_count += 1

        except Exception as e:
            print(f"  错误: {e}")
            import traceback

            traceback.print_exc()

    wb.save(excel_path)
    print(
        f"\n最终: {ws.max_row} 行 (新增 {new_rows_count} 行, 跳过 {skipped_rows_count} 行)"
    )
    print(f"已保存到 {excel_path}")


if __name__ == "__main__":
    main()
