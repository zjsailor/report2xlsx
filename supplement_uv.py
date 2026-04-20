# -*- coding: utf-8 -*-
"""
补充U列和V列脚本
用于从MD文件中提取SNP分型一致性，补充到Excel中

用法:
    python supplement_u.py
"""
import openpyxl
import os
import re
from pathlib import Path


def get_empty_snp_rows():
    """获取U列或V列为空的行"""
    wb = openpyxl.load_workbook('Info.xlsx')
    ws = wb.active
    empty_rows = []
    for row in range(2, ws.max_row + 1):
        filename = ws.cell(row, 1).value
        embryo_id = ws.cell(row, 13).value
        u_value = ws.cell(row, 21).value
        v_value = ws.cell(row, 22).value
        sample_barcode = ws.cell(row, 3).value
        gene = ws.cell(row, 10).value
        if u_value is None or u_value == '' or v_value is None or v_value == '':
            empty_rows.append({
                'row': row,
                'filename': filename,
                'embryo_id': embryo_id,
                'barcode': sample_barcode,
                'gene': gene,
                'current_u': u_value,
                'current_v': v_value
            })
    return wb, ws, empty_rows


def extract_snp_from_md(md_folder, filename_pattern):
    """从MD文件中提取SNP分型一致性"""
    md_files = list(Path(md_folder).glob('*.md'))
    for md_file in md_files:
        if filename_pattern.replace('.pdf', '.md') in md_file.name:
            with open(md_file, 'r', encoding='utf-8') as f:
                content = f.read()
            return extract_snp_from_content(content, md_file.name)
    return {}


def extract_snp_from_content(content, filename=''):
    """从MD内容中提取SNP分型一致性

    返回:
        genes: [gene1, gene2]
        snp_results: {embryo_id: {gene1: snp1, gene2: snp2}}
    """
    genes = []
    snp_results = {}

    lines = content.split('\n')
    in_target_section = False
    sample_name_col = -1

    for i, line in enumerate(lines):
        stripped = line.strip()

        if '目标变异检测结果' in stripped or '目标变异/SNP分型' in stripped:
            in_target_section = True
            continue

        if not in_target_section:
            continue

        if not stripped or stripped.startswith('|---'):
            continue

        if any(x in stripped for x in ['SNP可用位点', 'SNP单体型', '位点验证', '附件', '结果说明', '图谱']):
            break

        # 检测基因行
        gene_match = re.match(r'^([A-Z][A-Za-z0-9_]+),', stripped)
        if gene_match:
            gene_name = gene_match.group(1)
            if gene_name not in genes:
                genes.append(gene_name)
            continue

        if '样本名称' in stripped and '目标变异' in stripped:
            parts = [p.strip() for p in stripped.split('|') if p.strip()]
            for j, p in enumerate(parts):
                if '样本名称' in p:
                    sample_name_col = j
            continue

        # 处理带|的表格行
        if sample_name_col >= 0 and '|' in stripped:
            parts = [p.strip() for p in stripped.split('|') if p.strip()]
            if len(parts) > sample_name_col:
                sample_id = parts[sample_name_col]
                if sample_id and sample_id not in ['样本名称'] and not sample_id.startswith('基因'):
                    if sample_id not in snp_results:
                        snp_results[sample_id] = {}

                    for idx, gene in enumerate(genes):
                        col_offset = sample_name_col + 1 + idx * 2
                        snp_col_offset = col_offset + 1
                        if len(parts) > snp_col_offset:
                            for p in reversed(parts[snp_col_offset:]):
                                if p in ['一致', '不一致', '不一致（位点扩增ADO）', '不一致（位点扩增）', '不一致（检测异常）']:
                                    snp_results[sample_id][gene] = p
                                    break
                                if p == '-' or p == '—' or p == '':
                                    if gene not in snp_results[sample_id] or snp_results[sample_id].get(gene, '') == '':
                                        snp_results[sample_id][gene] = '不一致'
                                    break
                            else:
                                if gene not in snp_results[sample_id]:
                                    snp_results[sample_id][gene] = ''
            continue

        # 处理无|的文本行
        parts = stripped.split()
        if len(parts) >= 2:
            first_part = parts[0]
            if first_part and not first_part.startswith('基因') and not first_part.startswith('DMD') and not first_part.startswith('FBN'):
                if first_part not in ['SNP连锁分析判断结果', '与突变位点检测结果', '是否一致', '可用位点数']:
                    if first_part not in snp_results:
                        snp_results[first_part] = {}

                    for idx, gene in enumerate(genes):
                        if len(parts) > 2 + idx * 2:
                            p = parts[2 + idx * 2]
                            if p in ['一致', '不一致', '不一致（位点扩增ADO）', '不一致（位点扩增）', '不一致（检测异常）']:
                                snp_results[first_part][gene] = p
                            elif p == '-' or p == '—':
                                snp_results[first_part][gene] = '不一致'
                            elif gene not in snp_results[first_part]:
                                snp_results[first_part][gene] = ''

    return genes, snp_results


def find_matching_embryo(embryo_id, genes, snp_results):
    """在SNP结果中查找匹配的胚胎ID"""
    if not embryo_id or not snp_results:
        return {}

    clean_id = embryo_id.replace(' ', '').replace('_', '')

    # 精确匹配
    if embryo_id in snp_results:
        return snp_results[embryo_id]

    # 去除空格后匹配
    for key in snp_results.keys():
        key_clean = key.replace(' ', '').replace('_', '')
        if key_clean == clean_id:
            return snp_results[key]

    # 数字前缀匹配
    embryo_nums = re.findall(r'\d+', embryo_id)
    if embryo_nums:
        embryo_first_num = embryo_nums[0]
        for key in snp_results.keys():
            key_nums = re.findall(r'\d+', key)
            if key_nums and key_nums[0] == embryo_first_num:
                return snp_results[key]

    return {}


def main():
    md_folder = r'D:\md2excel\xh\markdown'

    print('读取Info.xlsx中U列/V列为空的记录...')
    wb, ws, empty_rows = get_empty_snp_rows()
    print(f'找到 {len(empty_rows)} 条U列/V列为空的记录\n')

    if not empty_rows:
        print('没有需要补充的记录')
        return

    updated_count = 0
    log_entries = []

    for item in empty_rows:
        row = item['row']
        filename = item['filename']
        embryo_id = item['embryo_id']
        barcode = item['barcode']
        gene = item['gene']

        if not filename:
            continue

        md_pattern = filename.replace('.pdf', '')
        genes, snp_results = extract_snp_from_md(md_folder, md_pattern)

        if snp_results:
            matched_snps = find_matching_embryo(embryo_id, genes, snp_results)
            if matched_snps:
                u_updated = False
                v_updated = False

                if genes:
                    gene1 = genes[0]
                    gene2 = genes[1] if len(genes) > 1 else None

                    current_u = ws.cell(row, 21).value
                    current_v = ws.cell(row, 22).value

                    if (current_u is None or current_u == '') and gene1 in matched_snps:
                        ws.cell(row, 21).value = matched_snps[gene1]
                        u_updated = True

                    if (current_v is None or current_v == '') and gene2 and gene2 in matched_snps:
                        ws.cell(row, 22).value = matched_snps[gene2]
                        v_updated = True

                if u_updated or v_updated:
                    updated_count += 1
                    log_entries.append({
                        'row': row,
                        'filename': filename,
                        'embryo_id': embryo_id,
                        'barcode': barcode,
                        'u_value': ws.cell(row, 21).value if u_updated else item['current_u'],
                        'v_value': ws.cell(row, 22).value if v_updated else item['current_v'],
                        'status': f"U更新:{u_updated}, V更新:{v_updated}"
                    })
                    print(f'[Row {row}] {embryo_id} -> U:{ws.cell(row, 21).value}, V:{ws.cell(row, 22).value}')
            else:
                log_entries.append({
                    'row': row,
                    'filename': filename,
                    'embryo_id': embryo_id,
                    'barcode': barcode,
                    'u_value': '',
                    'v_value': '',
                    'status': f'胚胎ID未匹配 (MD中SNP: {list(snp_results.keys())[:5]}...)'
                })
                print(f'[Row {row}] {embryo_id} -> 未匹配到SNP结果')
        else:
            log_entries.append({
                'row': row,
                'filename': filename,
                'embryo_id': embryo_id,
                'barcode': barcode,
                'u_value': '',
                'v_value': '',
                'status': '未找到对应MD文件'
            })
            print(f'[Row {row}] {embryo_id} -> 未找到MD文件')

    wb.save('Info.xlsx')
    print(f'\n保存完成，共补充 {updated_count} 条记录')

    # 生成LOG
    with open('supplement_uv_log.txt', 'w', encoding='utf-8') as f:
        f.write('=' * 80 + '\n')
        f.write('U列/V列补充LOG\n')
        f.write('=' * 80 + '\n\n')
        f.write(f'总空白记录数: {len(empty_rows)}\n')
        f.write(f'成功补充数: {updated_count}\n\n')

        f.write('-' * 80 + '\n')
        f.write('详细记录:\n')
        f.write('-' * 80 + '\n')
        for entry in log_entries:
            f.write(f"Row {entry['row']}: {entry['embryo_id']} | {entry['status']}\n")
            if entry['u_value']:
                f.write(f"  U={entry['u_value']}, V={entry['v_value']}\n")

    print(f'\nLOG已保存到: supplement_uv_log.txt')


if __name__ == '__main__':
    main()