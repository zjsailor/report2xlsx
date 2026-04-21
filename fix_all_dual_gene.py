# -*- coding: utf-8 -*-
"""
修复双基因报告的U列和V列
从MD文件中提取SNP分型一致性，补充到Excel中

用法:
    python fix_all_dual_gene.py -i <md文件夹> -o <输出excel>
    python fix_all_dual_gene.py -i ./markdown -o Info100smp.xlsx
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
import re
from pathlib import Path
from collections import defaultdict
import argparse


def extract_snp_from_md(filepath):
    """Extract SNP typing consistency for all genes from a markdown file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    genes = []
    snp_results = {}
    lines = content.split('\n')

    current_gene = None
    in_target_section = False

    for line in lines:
        stripped = line.strip()

        if '目标变异检测结果' in stripped:
            in_target_section = True
            continue

        if in_target_section and any(x in stripped for x in ['SNP可用位点', 'SNP单体型', '位点验证', '附件']):
            break

        if not in_target_section:
            continue

        if not stripped or stripped.startswith('|---'):
            continue

        gene_match = re.match(r'^([A-Z][A-Za-z0-9_]+),c\.', stripped)
        if gene_match:
            current_gene = gene_match.group(1)
            if current_gene not in genes:
                genes.append(current_gene)
            continue

        if '|' not in stripped and current_gene:
            parts = stripped.split()
            if len(parts) >= 2:
                first = parts[0]
                if len(first) < 20 and not first.startswith('-'):
                    embryo = first
                    last = parts[-1]

                    if embryo not in snp_results:
                        snp_results[embryo] = {}

                    if '不一致' in last:
                        if '不一致（位点扩增ADO）' in last:
                            snp_results[embryo][current_gene] = '不一致（位点扩增ADO）'
                        elif '不一致（' in last:
                            snp_results[embryo][current_gene] = last
                        else:
                            snp_results[embryo][current_gene] = '不一致'
                    elif '一致' in last:
                        snp_results[embryo][current_gene] = '一致'
                    elif last == '-' or last == '—':
                        snp_results[embryo][current_gene] = '不一致'
            continue

        if '|' in stripped:
            parts = [p.strip() for p in stripped.split('|') if p.strip()]
            if len(parts) >= 2:
                first = parts[0]
                if len(first) < 20 and not first.startswith('-') and first not in ['样本名称']:
                    embryo = first
                    if embryo not in snp_results:
                        snp_results[embryo] = {}

                    for p in reversed(parts):
                        if '不一致' in p:
                            if '不一致（位点扩增ADO）' in p:
                                snp_results[embryo][current_gene] = '不一致（位点扩增ADO）'
                            elif '不一致（' in p:
                                snp_results[embryo][current_gene] = p
                            else:
                                snp_results[embryo][current_gene] = '不一致'
                            break
                        elif '一致' in p:
                            snp_results[embryo][current_gene] = '一致'
                            break
                        elif p == '-' or p == '—':
                            snp_results[embryo][current_gene] = '不一致'
                            break

    return genes, snp_results


def find_matching_embryo(embryo_id, snp_results):
    """Find matching embryo ID in SNP results with fuzzy matching."""
    if not embryo_id or not snp_results:
        return {}

    clean_id = str(embryo_id).strip()

    if clean_id in snp_results:
        return snp_results[clean_id]

    clean_id_no_star = clean_id.replace('*', '')
    for key in snp_results.keys():
        key_no_star = str(key).replace('*', '')
        if key_no_star == clean_id_no_star:
            return snp_results[key]

    nums = re.findall(r'\d+', clean_id)
    if nums:
        for key in snp_results.keys():
            key_nums = re.findall(r'\d+', str(key))
            if key_nums and key_nums[0] == nums[0]:
                return snp_results[key]

    return {}


def main():
    parser = argparse.ArgumentParser(description='修复双基因报告的U列和V列')
    parser.add_argument('-i', '--input', required=True, help='MD文件夹路径')
    parser.add_argument('-o', '--output', required=True, help='输出Excel文件路径')
    args = parser.parse_args()

    md_folder = args.input
    excel_file = args.output

    print(f'输入文件夹: {md_folder}')
    print(f'输出Excel: {excel_file}')

    md_path = Path(md_folder)
    if not md_path.exists():
        print(f'错误: 文件夹不存在 {md_folder}')
        return

    md_files = list(md_path.glob('*.md'))
    print(f'找到 {len(md_files)} 个MD文件')

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    rows_data = []
    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
        row_values = [cell.value for cell in row]
        rows_data.append({
            'row_num': i,
            'filename': row_values[0],
            'embryo_id': row_values[12],
            'gene': row_values[9],
            'u_val': row_values[20] if len(row_values) > 20 else None,
            'v_val': row_values[21] if len(row_values) > 21 else None
        })

    by_file = defaultdict(list)
    for r in rows_data:
        if r['gene'] and ',' in str(r['gene']):
            by_file[r['filename']].append(r)

    total_u_updated = 0
    total_v_updated = 0

    for filename, file_rows in sorted(by_file.items()):
        md_pattern = filename.replace('.pdf', '.md')
        md_file = None
        for mf in md_files:
            if md_pattern in mf.name:
                md_file = mf
                break

        if not md_file:
            print(f'MD file not found for: {filename}')
            continue

        genes, snp_results = extract_snp_from_md(md_file)
        print(f'\nProcessing: {filename}')
        print(f'  Genes: {genes}')
        print(f'  Embryos found: {list(snp_results.keys())}')

        for r in file_rows:
            row_num = r['row_num']
            embryo_id = r['embryo_id']

            matched = find_matching_embryo(embryo_id, snp_results)
            if not matched:
                print(f'  WARNING: No match for embryo {embryo_id}')
                continue

            if len(genes) >= 1:
                gene1 = genes[0]
                if gene1 in matched and matched[gene1]:
                    old_u = ws.cell(row_num, 21).value
                    new_u = matched[gene1]
                    if old_u != new_u:
                        ws.cell(row_num, 21).value = new_u
                        print(f'  Row {row_num} (胚胎{embryo_id}): U changed from {old_u} to {new_u}')
                        total_u_updated += 1

            if len(genes) >= 2:
                gene2 = genes[1]
                if gene2 in matched and matched[gene2]:
                    old_v = ws.cell(row_num, 22).value
                    new_v = matched[gene2]
                    if old_v != new_v:
                        ws.cell(row_num, 22).value = new_v
                        print(f'  Row {row_num} (胚胎{embryo_id}): V changed from {old_v} to {new_v}')
                        total_v_updated += 1

    wb.save(excel_file)
    print('\n' + '='*60)
    print(f'Saved to: {excel_file}')
    print(f'Total U column updates: {total_u_updated}')
    print(f'Total V column updates: {total_v_updated}')

    wb2 = openpyxl.load_workbook(excel_file)
    ws2 = wb2.active
    dual_u_empty = 0
    dual_v_empty = 0
    for row in ws2.iter_rows(min_row=2):
        gene = row[9].value
        if gene and ',' in str(gene):
            if row[20].value is None:
                dual_u_empty += 1
            if row[21].value is None:
                dual_v_empty += 1
    print(f'Remaining empty U: {dual_u_empty}')
    print(f'Remaining empty V: {dual_v_empty}')


if __name__ == '__main__':
    main()